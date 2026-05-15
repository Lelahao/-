"""人员读取与 CRUD / 导入（GET detail 同结构；写操作单独 REST）。"""

from __future__ import annotations

import io
from typing import Any

from fastapi import APIRouter, File, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, field_validator

from server.api.common import run_read, run_write
from server.repositories import people_repo, plan_repo

router = APIRouter(prefix="/api", tags=["people"])


class PeopleBatchBody(BaseModel):
    people: list[dict[str, Any]]
    replace: bool = False


class CreatePersonBody(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    name: str
    region: str = ""
    position: str = ""
    role: str = ""

    @field_validator("name")
    @classmethod
    def name_nonempty(cls, v: str) -> str:
        s = str(v).strip()
        if not s:
            raise ValueError("cannot be empty")
        return s

    @field_validator("region", "position", "role")
    @classmethod
    def optional_strip(cls, v: str) -> str:
        return str(v).strip()


class UpdatePersonBody(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    name: str | None = Field(default=None, alias="name")
    region: str | None = None
    position: str | None = None
    role: str | None = None

    @field_validator("name")
    @classmethod
    def name_strip_nonempty(cls, v: str | None) -> str | None:
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            raise ValueError("cannot be empty")
        return s

    @field_validator("region", "position", "role")
    @classmethod
    def optional_strip_when_present(cls, v: str | None) -> str | None:
        if v is None:
            return None
        return str(v).strip()


def _parse_people_xlsx(data: bytes) -> list[tuple[str, str, str, str]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise ValueError("empty sheet")
        header_map: dict[str, int] = {}
        aliases = {
            "name": ("姓名", "name", "Name"),
            "region": ("区域", "region", "Region"),
            "position": ("岗位", "position", "Position"),
            "role": ("角色", "role", "Role"),
        }
        for key, labels in aliases.items():
            for i, h in enumerate(header):
                if h is None:
                    continue
                hs = str(h).strip()
                if hs in labels:
                    header_map[key] = i
                    break
        if "name" not in header_map:
            raise ValueError("missing column: name")

        def cell(row: tuple[Any, ...] | None, key: str) -> str:
            if row is None or key not in header_map:
                return ""
            idx = header_map[key]
            if idx >= len(row):
                return ""
            v = row[idx]
            return "" if v is None else str(v).strip()

        out: list[tuple[str, str, str, str]] = []
        for row in rows_iter:
            if row is None:
                continue

            name = cell(row, "name")
            region = cell(row, "region")
            position = cell(row, "position")
            role = cell(row, "role")
            if not name and not region and not position and not role:
                continue
            out.append((name, region, position, role))
        return out
    finally:
        wb.close()


@router.get("/plans/{plan_id}/people")
def list_people(plan_id: str):
    def r(conn):
        d = plan_repo.get_plan_detail(conn, plan_id)
        return {"people": d["people"]}

    return run_read(r)


@router.put("/plans/{plan_id}/people")
def put_people(plan_id: str, body: PeopleBatchBody):
    def w(conn):
        row = conn.execute(
            "SELECT 1 FROM plans WHERE id = ?1",
            (plan_id,),
        ).fetchone()
        if row is None:
            raise ValueError("plan not found")
        return {
            "planUpdatedAt": people_repo.upsert_people(
                conn, plan_id, body.people, replace=body.replace
            ),
        }

    return run_write(w)


@router.post("/plans/{plan_id}/people/import")
async def import_people(plan_id: str, file: UploadFile = File(...)):
    raw = await file.read()
    if not raw:
        raise ValueError("empty file")
    lower = (file.filename or "").lower()
    if not lower.endswith(".xlsx"):
        raise ValueError("only .xlsx supported")

    try:
        rows = _parse_people_xlsx(raw)
    except Exception as e:
        raise ValueError(str(e) or "invalid xlsx") from e

    def w(conn):
        row = conn.execute(
            "SELECT 1 FROM plans WHERE id = ?1",
            (plan_id,),
        ).fetchone()
        if row is None:
            raise ValueError("plan not found")
        return people_repo.import_people_rows(conn, plan_id, rows)

    return run_write(w)


@router.post("/plans/{plan_id}/people")
def create_person_route(plan_id: str, body: CreatePersonBody):
    def w(conn):
        person, plan_updated = people_repo.create_person(
            conn,
            plan_id,
            display_name=body.name,
            region=body.region,
            position=body.position,
            role=body.role,
        )
        return {"person": person, "planUpdatedAt": plan_updated}

    return run_write(w)


@router.patch("/plans/{plan_id}/people/{person_id}")
def update_person_route(plan_id: str, person_id: str, body: UpdatePersonBody):
    def w(conn):
        if all(
            x is None
            for x in (body.name, body.region, body.position, body.role)
        ):
            raise ValueError("no fields to update")
        person, plan_updated = people_repo.update_person(
            conn,
            plan_id,
            person_id,
            display_name=body.name,
            region=body.region,
            position=body.position,
            role=body.role,
        )
        return {"person": person, "planUpdatedAt": plan_updated}

    return run_write(w)


@router.delete("/plans/{plan_id}/people/{person_id}")
def delete_person_route(plan_id: str, person_id: str):
    def w(conn):
        plan_updated = people_repo.delete_person(conn, plan_id, person_id)
        return {"ok": True, "planUpdatedAt": plan_updated}

    return run_write(w)


@router.post("/plans/{plan_id}/people/{person_id}/unassign")
def unassign_person_route(plan_id: str, person_id: str):
    def w(conn):
        plan_updated = people_repo.unassign_person(conn, plan_id, person_id)
        return {"ok": True, "planUpdatedAt": plan_updated}

    return run_write(w)
